"""Parser XLSX.

Regla: si existe un CSV equivalente, se ignora el XLSX (el XLSX corrompe los IDs a
notación científica: `3.2634855E7` en vez de `32634855`). Eso lo decide `roles.py`.

Para los que sí se procesan, dos cuidados:
  · NUNCA confiar en `max_row`: `mag-conferences-list.xlsx` declara 999 filas XML y
    contiene 28 con datos (971 vacías);
  · castear los IDs numéricos a `int` o los PMIDs quedan corruptos.
"""

from __future__ import annotations

from pathlib import Path

from ...normalization.unicode_clean import normalize_value
from ..base import BlockBuilder, ParseResult
from .csv_parser import row_to_text

# Un valor float que es en realidad un identificador entero. openpyxl devuelve 32634855.0
# para `3.2634855E7`; sin este casteo el PMID se persiste como "32634855.0".
_ID_HINTS = ("id", "pmid", "pmcid", "doi", "code", "codigo", "number", "año", "year")


def _coerce(name: str, value: object) -> object:
    if value is None:
        return None
    if isinstance(value, float):
        if value.is_integer():
            lowered = name.lower()
            if any(h in lowered for h in _ID_HINTS) or abs(value) >= 1e6:
                return int(value)
        return value
    if isinstance(value, str):
        return normalize_value(value) or None
    return value


def parse_xlsx(path: Path, doc_id: str, pipeline_version: str) -> ParseResult:
    from openpyxl import load_workbook

    result = ParseResult()
    builder = BlockBuilder(doc_id, pipeline_version)

    wb = load_workbook(path, read_only=True, data_only=True)
    sheets_meta: list[dict] = []

    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        header: list[str] = []
        # `ws.max_row` NO es fiable (y en modo read_only vale 0 hasta iterar): se cuentan
        # las filas realmente emitidas por el XML frente a las que traen datos.
        declared = 0
        data_rows = 0

        for raw in rows_iter:
            declared += 1
            values = list(raw)
            if not any(v is not None and str(v).strip() for v in values):
                continue  # fila vacía: 971 de las 999 de mag-conferences-list
            if not header:
                header = [normalize_value(str(v)) if v is not None else "" for v in values]
                continue

            data_rows += 1
            structured = {}
            for name, value in zip(header, values):
                coerced = _coerce(name, value)
                if coerced is not None and coerced != "":
                    structured[name or "?"] = coerced

            if not structured:
                continue

            text = row_to_text(header, [str(v) if v is not None else "" for v in values])
            if block := builder.add(
                text,
                "table_row",
                "structured",
                row=data_rows,
                structured_data=structured,
            ):
                result.blocks.append(block)

        sheets_meta.append(
            {"sheet": ws.title, "declared_rows": declared, "data_rows": data_rows}
        )
        # Tres XLSX son datasets abandonados a medias en el origen.
        if declared > 50 and data_rows and declared / max(data_rows, 1) > 5:
            result.warnings.append(
                f"abandoned_source:{ws.title}__{data_rows}_of_{declared}_declared_rows"
            )

    wb.close()
    result.metadata = {"schema": "xlsx_table", "sheets": sheets_meta}
    return result
